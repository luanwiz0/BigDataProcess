#!/usr/bin/python3
import openpyxl
import sys

def check_index(list, index):
	while(list[index] == list[index + 1]):
		if(index == -1):
			break
		index -= 1
	return index

wb = openpyxl.load_workbook(filename = "student.xlsx")
ws = wb['Sheet1']

total_list = []

row_id = 2
for row in ws:
	sum = ws.cell(row = row_id, column = 3).value
	if sum is None:
		break
	else:
		sum *= 0.3
	sum += ws.cell(row = row_id, column = 4).value * 0.35
	sum += ws.cell(row = row_id, column = 5).value * 0.34
	sum += ws.cell(row = row_id, column = 6).value
	ws.cell(row = row_id, column = 7).value = sum
	total_list.append(ws.cell(row = row_id, column = 7).value)
	row_id += 1


total_list.sort(reverse = True)
student_num = len(total_list)
a_index = int(student_num * 0.3) - 1
b_index = int(student_num * 0.7) - 1

a_index = check_index(total_list, a_index)
b_index = check_index(total_list, b_index)

a_plus_index = -1
b_plus_index = -1
c_plus_index = -1

if a_index != -1:
	a_plus_index = check_index(total_list, int(a_index * 0.5))
	if a_plus_index == 0 and a_index == 0:
		a_plus_index = -1
if b_index != -1:
	b_plus_index = check_index(total_list, int((a_index + b_index) * 0.5))
c_plus_index = check_index(total_list, int((len(total_list) + b_index) * 0.5))

row_id = 2
for row in ws:
	total = ws.cell(row = row_id, column = 7).value
	if total is None:
		break
	grade = "C0"
	if c_plus_index != -1:
		if total >= total_list[c_plus_index]:
			grade = "C+"
	if b_index != -1:
		if total >= total_list[b_index]:
			grade = "B0"
	if b_plus_index != -1:
		if total >= total_list[b_plus_index]:
			grade = "B+"
	if a_index != -1:
		if total >= total_list[a_index]:
			grade = "A0"
	if a_plus_index != -1:
		if total >= total_list[a_plus_index]:
			grade = "A+"
	ws.cell(row = row_id, column = 8).value = grade
	row_id += 1


wb.save("student.xlsx")
